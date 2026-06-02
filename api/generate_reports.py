"""Generate Excel and PDF reports for the RePrime Data Platform."""
import csv
import json
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data" / "scraped"
CSV_PATH = Path(__file__).parent / "data" / "sources_611.csv"
OUTPUT_DIR = Path(__file__).parent / "data" / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_sources() -> dict:
    """Load source index from CSV."""
    index = {}
    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            sid = row.get("ID", "").strip()
            if sid:
                index[sid] = {
                    "id": sid,
                    "name": row.get("Source Name", "").strip(),
                    "provider": row.get("Provider", "").strip(),
                    "category": row.get("Category", "").strip(),
                    "endpoint": row.get("Endpoint URL", "").strip(),
                    "auth": row.get("Auth Required", "none").strip(),
                    "cost": row.get("Monthly Cost (USD)", "0").strip().replace("$", "").replace(",", ""),
                    "source_type": row.get("Source Type", "").strip(),
                }
    return index


def analyze_scraped() -> tuple:
    """Analyze all scraped files."""
    results = []
    for path in sorted(DATA_DIR.glob("*.json")):
        with open(path) as f:
            data = json.load(f)
        sid = path.stem
        rc = data.get("record_count", 0)
        recs = data.get("records", [])
        is_error = rc == 1 and len(recs) == 1 and "error_code" in (recs[0] if recs else {})
        results.append({
            "source_id": sid,
            "scraped_at": data.get("scraped_at", ""),
            "record_count": 0 if is_error else rc,
            "status": "error" if is_error else "success",
            "error": recs[0].get("error_message", "") if is_error else "",
        })
    return results


def generate_excel(sources: dict, scraped: list) -> str:
    """Generate Excel report using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, BarChart, Reference
    except ImportError:
        # Fallback: generate CSV
        return generate_csv_report(sources, scraped)

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    title_font = Font(size=16, bold=True, color="1E3A5F")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    red_fill = PatternFill("solid", fgColor="F8D7DA")

    ws.merge_cells("A1:F1")
    ws["A1"] = "RePrime Data Platform — Scrape Report"
    ws["A1"].font = title_font
    ws["A3"] = "Generated:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Total Sources:"
    ws["B4"] = len(sources)

    scraped_map = {s["source_id"]: s for s in scraped}
    success_count = sum(1 for s in scraped if s["status"] == "success")
    error_count = sum(1 for s in scraped if s["status"] == "error")
    total_records = sum(s["record_count"] for s in scraped)
    paid_count = sum(1 for s in sources.values() if float(s.get("cost") or 0) > 0)

    ws["A5"] = "Scraped (success):"
    ws["B5"] = success_count
    ws["A6"] = "Errors:"
    ws["B6"] = error_count
    ws["A7"] = "Paid (skipped):"
    ws["B7"] = paid_count
    ws["A8"] = "Total Records:"
    ws["B8"] = total_records
    ws["A9"] = "Success Rate:"
    ws["B9"] = f"{success_count / max(len(sources) - paid_count, 1) * 100:.1f}%"

    # Category summary
    by_cat = defaultdict(lambda: {"total": 0, "scraped": 0, "records": 0})
    for sid, meta in sources.items():
        cat = meta.get("category", "other") or "other"
        by_cat[cat]["total"] += 1
        if sid in scraped_map and scraped_map[sid]["status"] == "success":
            by_cat[cat]["scraped"] += 1
            by_cat[cat]["records"] += scraped_map[sid]["record_count"]

    row = 12
    ws.cell(row=row, column=1, value="Category").font = header_font
    ws.cell(row=row, column=2, value="Total").font = header_font
    ws.cell(row=row, column=3, value="Scraped").font = header_font
    ws.cell(row=row, column=4, value="Records").font = header_font
    ws.cell(row=row, column=5, value="Rate").font = header_font
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = header_fill

    for cat, vals in sorted(by_cat.items(), key=lambda x: -x[1]["total"]):
        row += 1
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=vals["total"])
        ws.cell(row=row, column=3, value=vals["scraped"])
        ws.cell(row=row, column=4, value=vals["records"])
        rate = vals["scraped"] / max(vals["total"], 1) * 100
        ws.cell(row=row, column=5, value=f"{rate:.0f}%")

    # --- Sheet 2: All Sources ---
    ws2 = wb.create_sheet("All Sources")
    headers = ["ID", "Name", "Provider", "Category", "Auth", "Type", "Cost", "Status", "Records", "Scraped At"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for sid, meta in sorted(sources.items()):
        s = scraped_map.get(sid)
        cost_val = float(meta.get("cost") or 0)
        if cost_val > 0:
            status = "paid_skipped"
        elif s:
            status = s["status"]
        else:
            status = "not_scraped"

        ws2.cell(row=row, column=1, value=sid)
        ws2.cell(row=row, column=2, value=meta.get("name", ""))
        ws2.cell(row=row, column=3, value=meta.get("provider", ""))
        ws2.cell(row=row, column=4, value=meta.get("category", ""))
        ws2.cell(row=row, column=5, value=meta.get("auth", ""))
        ws2.cell(row=row, column=6, value=meta.get("source_type", ""))
        ws2.cell(row=row, column=7, value=cost_val)
        cell = ws2.cell(row=row, column=8, value=status)
        if status == "success":
            cell.fill = green_fill
        elif status == "error":
            cell.fill = red_fill
        ws2.cell(row=row, column=9, value=s["record_count"] if s else 0)
        ws2.cell(row=row, column=10, value=s["scraped_at"] if s else "")
        row += 1

    for col in ws2.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # --- Sheet 3: Errors ---
    ws3 = wb.create_sheet("Errors")
    err_headers = ["Source ID", "Name", "Error"]
    for c, h in enumerate(err_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for s in scraped:
        if s["status"] == "error":
            ws3.cell(row=row, column=1, value=s["source_id"])
            ws3.cell(row=row, column=2, value=sources.get(s["source_id"], {}).get("name", ""))
            ws3.cell(row=row, column=3, value=s["error"][:200])
            row += 1

    out_path = OUTPUT_DIR / f"reprime_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(out_path))
    return str(out_path)


def generate_csv_report(sources: dict, scraped: list) -> str:
    """Fallback CSV report."""
    scraped_map = {s["source_id"]: s for s in scraped}
    out_path = OUTPUT_DIR / f"reprime_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Name", "Provider", "Category", "Auth", "Status", "Records"])
        for sid, meta in sorted(sources.items()):
            s = scraped_map.get(sid)
            status = s["status"] if s else "not_scraped"
            w.writerow([sid, meta.get("name"), meta.get("provider"), meta.get("category"),
                        meta.get("auth"), status, s["record_count"] if s else 0])
    return str(out_path)


def generate_pdf(sources: dict, scraped: list) -> str:
    """Generate a PDF report as HTML-to-text summary."""
    scraped_map = {s["source_id"]: s for s in scraped}
    success_count = sum(1 for s in scraped if s["status"] == "success")
    error_count = sum(1 for s in scraped if s["status"] == "error")
    total_records = sum(s["record_count"] for s in scraped)
    paid_count = sum(1 for s in sources.values() if float(s.get("cost") or 0) > 0)

    by_cat = defaultdict(lambda: {"total": 0, "scraped": 0, "records": 0})
    for sid, meta in sources.items():
        cat = meta.get("category", "other") or "other"
        by_cat[cat]["total"] += 1
        if sid in scraped_map and scraped_map[sid]["status"] == "success":
            by_cat[cat]["scraped"] += 1
            by_cat[cat]["records"] += scraped_map[sid]["record_count"]

    lines = []
    lines.append("=" * 70)
    lines.append("REPRIME DATA PLATFORM — SCRAPE REPORT")
    lines.append("=" * 70)
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("EXECUTIVE SUMMARY")
    lines.append("-" * 40)
    lines.append(f"Total Sources:        {len(sources)}")
    lines.append(f"Scraped (success):    {success_count}")
    lines.append(f"Errors:               {error_count}")
    lines.append(f"Paid (skipped):       {paid_count}")
    lines.append(f"Total Records:        {total_records:,}")
    rate = success_count / max(len(sources) - paid_count, 1) * 100
    lines.append(f"Success Rate:         {rate:.1f}%")
    lines.append("")
    lines.append("CATEGORY BREAKDOWN")
    lines.append("-" * 40)
    lines.append(f"{'Category':<25} {'Total':>6} {'Scraped':>8} {'Records':>8} {'Rate':>6}")
    lines.append("-" * 55)
    for cat, vals in sorted(by_cat.items(), key=lambda x: -x[1]["total"]):
        r = vals["scraped"] / max(vals["total"], 1) * 100
        lines.append(f"{cat:<25} {vals['total']:>6} {vals['scraped']:>8} {vals['records']:>8} {r:>5.0f}%")

    lines.append("")
    lines.append("TOP 20 SOURCES BY RECORD COUNT")
    lines.append("-" * 40)
    top = sorted(scraped, key=lambda x: -x["record_count"])[:20]
    for i, s in enumerate(top, 1):
        name = sources.get(s["source_id"], {}).get("name", "Unknown")[:40]
        lines.append(f"{i:>2}. {s['source_id']} — {name} ({s['record_count']:,} records)")

    lines.append("")
    lines.append("ERRORS")
    lines.append("-" * 40)
    errors = [s for s in scraped if s["status"] == "error"]
    for s in errors:
        name = sources.get(s["source_id"], {}).get("name", "Unknown")[:40]
        lines.append(f"  {s['source_id']} — {name}: {s['error'][:80]}")

    lines.append("")
    lines.append("METHODOLOGY")
    lines.append("-" * 40)
    lines.append("1. Sources loaded from master CSV (611 entries)")
    lines.append("2. 9 paid APIs excluded (total ~$376/mo)")
    lines.append("3. 602 free sources attempted via async scraping")
    lines.append("4. Connectors: REST API, RSS/Atom, Socrata SODA, ArcGIS, Bulk Download")
    lines.append("5. Concurrency: 20 simultaneous connections via asyncio.Semaphore")
    lines.append("6. Rate limiting: Per-source configurable, default 60 req/min")
    lines.append("")
    lines.append("TECH STACK")
    lines.append("-" * 40)
    lines.append("Python 3.12+ | httpx | asyncio | FastAPI | feedparser")
    lines.append("Data stored as JSON files in api/data/scraped/")
    lines.append("")
    lines.append("=" * 70)
    lines.append("Built by RePrime Data Platform v1.0")
    lines.append("=" * 70)

    out_path = OUTPUT_DIR / f"reprime_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    with open(out_path, "w") as f:
        f.write("\n".join(lines))
    return str(out_path)


if __name__ == "__main__":
    print("Loading sources...")
    sources = load_sources()
    print(f"Loaded {len(sources)} sources from CSV")

    print("Analyzing scraped data...")
    scraped = analyze_scraped()
    print(f"Found {len(scraped)} scraped files")

    print("Generating Excel report...")
    xlsx_path = generate_excel(sources, scraped)
    print(f"Excel: {xlsx_path}")

    print("Generating text report...")
    txt_path = generate_pdf(sources, scraped)
    print(f"Report: {txt_path}")

    print("Done!")
